"""Build a multi-sheet Excel workbook from a computed plan result."""

import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.drawing.image import Image as XLImage

from .charts import year_trend_png, staffing_curve_png


def build_workbook(result: dict, plan: dict, explanation: str | None = None,
                   review: str | None = None) -> bytes:
    wb = Workbook()
    bold = Font(bold=True)
    head_fill = PatternFill("solid", fgColor="E8ECFB")

    def header(ws, row, cols):
        for c, val in enumerate(cols, 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.font = bold
            cell.fill = head_fill

    # ---- Summary sheet ----
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "WFM Holiday Plan — Summary"
    ws["A1"].font = Font(bold=True, size=14)
    holiday = plan.get("holiday") or {}
    q = plan.get("queue") or {}
    ws["A2"] = "Queue"; ws["B2"] = f"{q.get('name', '—')} ({q.get('region', '—')} / {q.get('type', '—')})"
    ws["A3"] = "Holiday"; ws["B3"] = holiday.get("name", "—")
    ws["A4"] = "Plan date"; ws["B4"] = holiday.get("plan_date", "—")
    ws["A5"] = "Total shrinkage %"; ws["B5"] = result["shrinkage"]["total_shrinkage_pct"]

    header(ws, 7, ["Channel", "Recommended combo", "Weekly forecast", "Busiest day", "Peak Net HC", "Peak Gross HC"])
    r = 8
    for ch, res in result["channels"].items():
        ws.cell(row=r, column=1, value=ch)
        ws.cell(row=r, column=2, value=(res["recommended"] or {}).get("combo", "—"))
        ws.cell(row=r, column=3, value=res["weekly_forecast"])
        ws.cell(row=r, column=4, value=res["busiest_day"])
        ws.cell(row=r, column=5, value=res["peak_net_hc"])
        ws.cell(row=r, column=6, value=res["peak_gross_hc"])
        r += 1

    # ---- Per-channel sheets ----
    for ch, res in result["channels"].items():
        s = wb.create_sheet(title=f"{ch[:25]} HC")
        header(s, 1, ["Slot start", "Slot end", "Calls/hr", "Net HC", "Gross HC"])
        rr = 2
        for iv in res["intervals"]:
            s.cell(row=rr, column=1, value=iv["start"])
            s.cell(row=rr, column=2, value=iv["end"])
            s.cell(row=rr, column=3, value=iv["calls_per_hour"])
            s.cell(row=rr, column=4, value=iv["net_hc"])
            s.cell(row=rr, column=5, value=iv["gross_hc"])
            rr += 1
        if not res["intervals"]:
            s.cell(row=2, column=1, value="(email = flat daily figure)")
            s.cell(row=2, column=4, value=res["peak_net_hc"])
            s.cell(row=2, column=5, value=res["peak_gross_hc"])

        # Embed the diagrams (charts) into the sheet
        try:
            yt = year_trend_png(res, ch)
            if yt:
                s.add_image(XLImage(io.BytesIO(yt)), "H1")
            sc = staffing_curve_png(res, ch)
            if sc:
                s.add_image(XLImage(io.BytesIO(sc)), "H16")
        except Exception:
            pass  # charts are presentation-only; never block the data export

        # combinations block
        cstart = rr + 2
        header(s, cstart, ["Combination", "Impact %", "Forecast", "Score", "Recommended"])
        for j, c in enumerate(res["combinations"], 1):
            s.cell(row=cstart + j, column=1, value=c["combo"])
            s.cell(row=cstart + j, column=2, value=c["blended_impact_pct"])
            s.cell(row=cstart + j, column=3, value=c["forecasted_volume"])
            s.cell(row=cstart + j, column=4, value=c["score"])
            s.cell(row=cstart + j, column=5, value="★" if c["recommended"] else "")

    # ---- AI Analysis sheet ----
    if explanation or review:
        ai = wb.create_sheet(title="AI Analysis")
        ai["A1"] = "AI Analysis (explains & advises only — never computes numbers)"
        ai["A1"].font = Font(bold=True, size=12)
        r = 3
        if explanation:
            ai.cell(row=r, column=1, value="Plan summary").font = bold
            r += 1
            for line in explanation.split("\n"):
                if line.strip():
                    ai.cell(row=r, column=1, value=line.strip()); r += 1
            r += 1
        if review:
            ai.cell(row=r, column=1, value="Analyst review & suggestions").font = bold
            r += 1
            for line in review.split("\n"):
                if line.strip():
                    ai.cell(row=r, column=1, value=line.strip()); r += 1
        ai.column_dimensions["A"].width = 110

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
